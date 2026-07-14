from io import StringIO, BytesIO
import csv
import json
import traceback
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from typing import Optional
from collections import defaultdict
from fastapi import FastAPI, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from .config import ASSETS, DEFAULT_ASSETS, load_assets_into
from .db import (
    init_db, upsert_rows, get_history, upsert_price_rows, get_price_history,
    seed_assets_if_empty, get_all_assets, insert_asset, delete_asset,
)
from .services import collect, collect_prices, collect_live, validate_asset_tickers
from .metrics import periods_per_year, interval_label
from .analysis import exchange_stats, monthly_table, funding_price_correlation, hourly_heatmap

EXCHANGES = ('binance', 'okx', 'hyperliquid')


def ms_to_dt(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000, tz=timezone.utc).strftime('%Y-%m-%d %H:%M:%S')


async def refresh_assets_cache() -> None:
    rows = await get_all_assets()
    load_assets_into(ASSETS, rows)


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    await seed_assets_if_empty(DEFAULT_ASSETS)
    await refresh_assets_cache()
    yield


app = FastAPI(title='Metals Funding History', lifespan=lifespan)
app.mount('/static', StaticFiles(directory='static'), name='static')
templates = Jinja2Templates(directory='templates')


@app.get('/', response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        'index.html', {'request': request, 'assets': ASSETS}
    )


@app.get('/analysis', response_class=HTMLResponse)
async def analysis_page(request: Request):
    asset_labels_json = json.dumps({k: v['label'] for k, v in ASSETS.items()}, ensure_ascii=False)
    return templates.TemplateResponse(
        'analysis.html',
        {'request': request, 'assets': ASSETS, 'asset_labels_json': asset_labels_json}
    )


@app.get('/api/assets')
async def list_assets():
    return JSONResponse({'ok': True, 'assets': ASSETS})


@app.post('/api/assets')
async def add_asset(request: Request):
    body = await request.json()
    key = (body.get('key') or '').strip().upper()
    label = (body.get('label') or '').strip()
    okx = (body.get('okx') or '').strip() or None
    binance = (body.get('binance') or '').strip() or None
    hyperliquid_dex = (body.get('hyperliquid_dex') or '').strip() or None
    hyperliquid_coin = (body.get('hyperliquid_coin') or '').strip() or None

    if not key or not key.replace('_', '').isalnum():
        return JSONResponse({'ok': False, 'error': 'Тикер актива обязателен и должен быть буквенно-цифровым'}, status_code=400)
    if not label:
        return JSONResponse({'ok': False, 'error': 'Название актива обязательно'}, status_code=400)
    if key in ASSETS:
        return JSONResponse({'ok': False, 'error': f'Актив {key} уже существует'}, status_code=409)
    if not any([okx, binance, hyperliquid_coin]):
        return JSONResponse({'ok': False, 'error': 'Нужен хотя бы один тикер (OKX, Binance или Hyperliquid)'}, status_code=400)
    if hyperliquid_coin and not hyperliquid_dex:
        return JSONResponse({'ok': False, 'error': 'Для Hyperliquid нужно указать dex (напр. "xyz" или "HyperEVM")'}, status_code=400)

    try:
        checks = await validate_asset_tickers(okx, binance, hyperliquid_dex, hyperliquid_coin)
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': f'Ошибка проверки тикеров: {e}'}, status_code=502)

    failed = [ex for ex, ok in checks.items() if not ok]
    if failed:
        return JSONResponse({
            'ok': False,
            'error': f'Тикер не подтверждён биржей: {", ".join(failed)}',
            'checks': checks,
        }, status_code=422)

    await insert_asset(key, label, okx, binance, hyperliquid_dex, hyperliquid_coin)
    await refresh_assets_cache()
    return JSONResponse({'ok': True, 'asset': key, 'checks': checks})


@app.delete('/api/assets/{key}')
async def remove_asset(key: str):
    key = key.upper()
    if key not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    await delete_asset(key)
    await refresh_assets_cache()
    return JSONResponse({'ok': True, 'asset': key})


@app.post('/api/sync/{asset}')
async def sync(asset: str):
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        rows = await collect(asset)
        inserted = await upsert_rows(rows)
        price_rows = await collect_prices(asset)
        price_inserted = await upsert_price_rows(price_rows)
        return JSONResponse({
            'ok': True, 'asset': asset,
            'received': len(rows), 'new': inserted,
            'price_received': len(price_rows), 'price_new': price_inserted,
        })
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse({'ok': False, 'error': str(e), 'detail': tb}, status_code=500)


@app.get('/api/live/{asset}')
async def live(asset: str):
    """Живой снэпшот (не история): predicted funding + basis по всем трём биржам."""
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        data = await collect_live(asset)
        return JSONResponse({'ok': True, 'asset': asset, 'exchanges': data})
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/price-history/{asset}')
async def price_history(asset: str, exchange: Optional[str] = Query(None)):
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        rows = await get_price_history(asset)
        if exchange:
            rows = [r for r in rows if r['exchange'] == exchange.lower()]
        return JSONResponse({'ok': True, 'asset': asset, 'count': len(rows), 'rows': rows})
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/analysis/{exchange}/{asset}')
async def analysis(exchange: str, asset: str):
    """Единый эндпоинт для экрана анализа: биржа+актив -> вся сводка одним ответом."""
    asset = asset.upper()
    exchange = exchange.lower()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    if exchange not in EXCHANGES:
        return JSONResponse({'ok': False, 'error': 'Неизвестная биржа'}, status_code=404)
    try:
        all_rows = await get_history(asset)
        by_exchange: dict = defaultdict(list)
        for r in all_rows:
            by_exchange[r['exchange']].append(r)

        rows = by_exchange.get(exchange, [])
        if not rows:
            return JSONResponse({
                'ok': True, 'exchange': exchange, 'asset': asset,
                'message': 'Нет данных для этой пары. Нажми «Собрать» сначала.',
            })

        cross_exchange = {
            ex: exchange_stats(ex_rows)
            for ex, ex_rows in by_exchange.items() if ex_rows
        }

        price_rows = await get_price_history(asset)
        price_rows_ex = [p for p in price_rows if p['exchange'] == exchange]
        price_series = [{'ts': p['ts'], 'close': p['close']} for p in price_rows_ex]

        funding_series = [
            {
                'ts': r['funding_time'],
                'rate_pct': round((r['funding_rate'] if r['funding_rate'] is not None else 0.0) * 100, 6),
            }
            for r in rows
        ]

        stats = exchange_stats(rows)
        stats['correlation_price'] = funding_price_correlation(rows, price_rows_ex)

        return JSONResponse({
            'ok': True,
            'exchange': exchange,
            'asset': asset,
            'symbol': rows[0]['symbol'],
            'stats': stats,
            'monthly': monthly_table(rows),
            'heatmap': hourly_heatmap(rows),
            'cross_exchange': cross_exchange,
            'funding_series': funding_series,
            'price_series': price_series,
        })
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/history/{asset}')
async def history(asset: str, exchange: Optional[str] = Query(None)):
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        rows = await get_history(asset)
        if exchange:
            rows = [r for r in rows if r['exchange'] == exchange.lower()]
        return JSONResponse({'ok': True, 'asset': asset, 'count': len(rows), 'rows': rows})
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/stats/{asset}')
async def stats(asset: str):
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        rows = await get_history(asset)
        if not rows:
            return JSONResponse({'ok': True, 'asset': asset,
                                 'message': 'Нет данных. Нажми «Собрать» сначала.'})
        by_exchange: dict = defaultdict(list)
        by_exchange_ts: dict = defaultdict(list)
        for r in rows:
            by_exchange[r['exchange']].append(r['funding_rate'])
            by_exchange_ts[r['exchange']].append(r['funding_time'])
        result = {'ok': True, 'asset': asset, 'total_rows': len(rows), 'by_exchange': {}}
        for exchange, rates in by_exchange.items():
            n = len(rates)
            avg = sum(rates) / n
            ppy = periods_per_year(by_exchange_ts[exchange])
            annual = avg * ppy * 100
            pos = sum(1 for r in rates if r > 0)
            neg = sum(1 for r in rates if r < 0)
            result['by_exchange'][exchange] = {
                'periods':          n,
                'periods_per_year': round(ppy, 1),
                'funding_interval': interval_label(ppy),
                'avg_rate_pct':     f'{avg * 100:.6f}%',
                'annualized_yield': f'{annual:.4f}%',
                'max_rate_pct':     f'{max(rates) * 100:.6f}%',
                'min_rate_pct':     f'{min(rates) * 100:.6f}%',
                'positive_periods': pos,
                'negative_periods': neg,
                'pct_positive':     f'{pos / n * 100:.1f}%',
            }
        return JSONResponse(result)
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/export/{asset}.csv')
async def export_csv(asset: str, exchange: Optional[str] = Query(None)):
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        rows = await get_history(asset)
        if exchange:
            rows = [r for r in rows if r['exchange'] == exchange.lower()]
        buf = StringIO()
        fieldnames = [
            'exchange', 'asset', 'symbol',
            'datetime_utc', 'funding_time_ms',
            'funding_rate', 'funding_rate_pct',
            'mark_price', 'premium',
        ]
        w = csv.DictWriter(buf, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            fr = row['funding_rate'] if row['funding_rate'] is not None else 0.0
            w.writerow({
                'exchange':         row['exchange'],
                'asset':            row['asset'],
                'symbol':           row['symbol'],
                'datetime_utc':     ms_to_dt(row['funding_time']),
                'funding_time_ms':  row['funding_time'],
                'funding_rate':     f'{fr:.10f}',
                'funding_rate_pct': f'{fr * 100:.8f}%',
                'mark_price':       row['mark_price'] if row['mark_price'] is not None else '',
                'premium':          f"{row['premium']:.10f}" if row['premium'] is not None else '',
            })
        buf.seek(0)
        suffix = f'_{exchange}' if exchange else ''
        fname = f'{asset.lower()}_funding{suffix}.csv'
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={fname}'}
        )
    except Exception as e:
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/export/{asset}.xlsx')
async def export_xlsx(asset: str, exchange: Optional[str] = Query(None)):
    """Excel-выгрузка с тремя листами:
       Raw — сырые данные
       Annual — годовая доходность по биржам (формула: СРЕДНЯЯ × ПЕРИОДОВ в ГОД)
       Monthly — помесячная доходность по биржам
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'ok': False, 'error': 'Неизвестный актив'}, status_code=404)
    try:
        rows = await get_history(asset)
        if exchange:
            rows = [r for r in rows if r['exchange'] == exchange.lower()]

        wb = openpyxl.Workbook()
        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(color='FFFFFF', bold=True)

        # ===== Лист 1: Raw =====
        ws_raw = wb.active
        ws_raw.title = 'Raw'
        raw_headers = [
            'exchange', 'asset', 'symbol', 'datetime_utc',
            'funding_time_ms', 'funding_rate', 'funding_rate_pct',
            'mark_price', 'premium'
        ]
        for col, h in enumerate(raw_headers, 1):
            cell = ws_raw.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            fr = row['funding_rate'] if row['funding_rate'] is not None else 0.0
            ws_raw.append([
                row['exchange'], row['asset'], row['symbol'],
                ms_to_dt(row['funding_time']), row['funding_time'],
                round(fr, 10), round(fr * 100, 8),
                row['mark_price'] if row['mark_price'] is not None else None,
                round(row['premium'], 10) if row['premium'] is not None else None,
            ])
        for col in range(1, len(raw_headers) + 1):
            ws_raw.column_dimensions[get_column_letter(col)].width = 20

        # ===== Лист 2: Annual =====
        # Формула: Годовая = Средняя ставка × Количество периодов в год (APR)
        ws_annual = wb.create_sheet('Annual')
        annual_headers = [
            'Биржа',
            'Количество периодов',
            'Интервал начисления',
            'Дата начала',
            'Дата окончания',
            'Дней наблюдений',
            'Средняя ставка, %',
            'Годовая доходность, %',
            'Макс ставка, %',
            'Мин ставка, %',
            '% положительных периодов',
        ]
        for col, h in enumerate(annual_headers, 1):
            cell = ws_annual.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        by_exchange: dict = defaultdict(list)
        by_exchange_ts: dict = defaultdict(list)
        for r in rows:
            ex = r['exchange']
            fr = r['funding_rate'] if r['funding_rate'] is not None else 0.0
            by_exchange[ex].append(fr)
            by_exchange_ts[ex].append(r['funding_time'])

        for ex, rates in by_exchange.items():
            n = len(rates)
            avg = sum(rates) / n
            # Правильная формула APR: средняя ставка × периодов в год.
            # Периоды в год считаем по факту данных, а не хардкодом по бирже,
            # т.к. интервал начисления зависит от конкретного инструмента
            # (напр. на Binance большинство товарных активов — 4ч, BTC — 8ч).
            timestamps = by_exchange_ts[ex]
            ppy = periods_per_year(timestamps)
            annual = avg * ppy * 100
            ts_min, ts_max = min(timestamps), max(timestamps)
            days = (ts_max - ts_min) / 86400000
            pos = sum(1 for r in rates if r > 0)
            ws_annual.append([
                ex,
                n,
                interval_label(ppy),
                ms_to_dt(ts_min),
                ms_to_dt(ts_max),
                round(days, 2),
                round(avg * 100, 6),
                round(annual, 4),
                round(max(rates) * 100, 6),
                round(min(rates) * 100, 6),
                round(pos / n * 100, 1),
            ])
        for col in range(1, len(annual_headers) + 1):
            ws_annual.column_dimensions[get_column_letter(col)].width = 26

        # ===== Лист 3: Monthly =====
        ws_monthly = wb.create_sheet('Monthly')
        monthly_headers = [
            'Биржа',
            'Месяц',
            'Количество периодов',
            'Средняя ставка за период, %',
            'Доходность за месяц, %',
            '% положительных периодов',
        ]
        for col, h in enumerate(monthly_headers, 1):
            cell = ws_monthly.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        monthly_data: dict = defaultdict(list)
        for r in rows:
            ex = r['exchange']
            fr = r['funding_rate'] if r['funding_rate'] is not None else 0.0
            dt = datetime.fromtimestamp(r['funding_time'] / 1000, tz=timezone.utc)
            monthly_data[(ex, dt.year, dt.month)].append(fr)

        for (ex, year, month), rates in sorted(monthly_data.items()):
            n = len(rates)
            s = sum(rates)
            avg = s / n
            pos = sum(1 for r in rates if r > 0)
            ws_monthly.append([
                ex,
                f'{year}-{month:02d}',
                n,
                round(avg * 100, 6),       # средняя ставка за период
                round(s * 100, 4),          # сумма всех ставок за месяц = доходность
                round(pos / n * 100, 1),
            ])
        for col in range(1, len(monthly_headers) + 1):
            ws_monthly.column_dimensions[get_column_letter(col)].width = 28

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        suffix = f'_{exchange}' if exchange else ''
        fname = f'{asset.lower()}_funding{suffix}.xlsx'
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={fname}'}
        )
    except Exception as e:
        print(traceback.format_exc())
        return JSONResponse({'ok': False, 'error': str(e)}, status_code=500)


@app.get('/api/debug/{asset}')
async def debug(asset: str):
    asset = asset.upper()
    if asset not in ASSETS:
        return JSONResponse({'error': 'Неизвестный актив'}, status_code=404)
    import httpx as hx
    results = {}
    now = datetime.now(tz=timezone.utc)
    start_ms = int(datetime(now.year, 1, 1, tzinfo=timezone.utc).timestamp() * 1000)
    cfg = ASSETS[asset]

    if cfg.get('okx'):
        try:
            async with hx.AsyncClient(timeout=15.0) as c:
                r = await c.get('https://www.okx.com/api/v5/public/funding-rate-history',
                                params={'instId': cfg['okx'], 'limit': 2})
                results['okx'] = {'status': r.status_code, 'instId': cfg['okx'],
                                  'sample': r.json().get('data', [])[:2]}
        except Exception as e:
            results['okx'] = {'error': str(e)}

    if cfg.get('binance'):
        try:
            async with hx.AsyncClient(timeout=15.0) as c:
                r = await c.get('https://fapi.binance.com/fapi/v1/fundingRate',
                                params={'symbol': cfg['binance'], 'limit': 2,
                                        'startTime': start_ms})
                results['binance'] = {'status': r.status_code, 'symbol': cfg['binance'],
                                      'sample': r.json()[:2] if r.status_code == 200 else r.text[:200]}
        except Exception as e:
            results['binance'] = {'error': str(e)}

    try:
        async with hx.AsyncClient(timeout=15.0) as c:
            r = await c.post('https://api.hyperliquid.xyz/info',
                             json={'type': 'fundingHistory',
                                   'coin': cfg['hyperliquid_coin'],
                                   'dex': cfg['hyperliquid_dex'],
                                   'startTime': start_ms,
                                   'endTime': start_ms + 86400000 * 3})
            results['hyperliquid'] = {
                'coin': cfg['hyperliquid_coin'], 'status': r.status_code,
                'sample': r.json()[:3] if r.status_code == 200 and isinstance(r.json(), list) else r.text[:300]
            }
    except Exception as e:
        results['hyperliquid'] = {'error': str(e)}

    return JSONResponse(results)
